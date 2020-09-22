import os
import sys
from os import path
# 便于计算宽度
import pandas as pd
import numpy as np
# 读写 xlsx 进行美化
from openpyxl import load_workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter, column_index_from_string

# 加根目录到 sys.path
project_folder = os.path.abspath(os.path.join(path.dirname(__file__), os.pardir))
if project_folder not in sys.path:
    sys.path.append(project_folder)

class xlsxfmt:
    """
    将所有 DataFrame 输出的 xlsx 文件进行格式化显示
    （适当的列宽、筛选、冻结首行、调整字体）
    """
    def __init__(self):
        self.folder = os.path.abspath(os.path.dirname(__file__))
    
    def format_all_xlsx_files(self):
        """
        把项目内，所有的 xlsx 文件都美化一下
        """
        # 获取所有 xlsx 文件
        xlsx_files = []
        for root, folders, files in os.walk(os.path.abspath('D:\github\deal-record-temp')):
            for f in files:
                if '.xlsx' in f:
                    xlsx_path = os.path.join(root, f)
                    xlsx_files.append(xlsx_path)
        # 逐一格式化
        for i, xlsx_file in enumerate(xlsx_files):
            print(f' [ {i + 1} / {len(xlsx_files)} ] 格式化 {xlsx_file}')
            self.format_xlsx_file(xlsx_file)
        pass
    
    def format_xlsx_file(self, xlsx_file):
        """
        格式化单个 xlsx 文件
        """
        # Step 0 加载数据，创建 DataFrame 和 Excel 对象
        wb = load_workbook(xlsx_file)
        ws = wb.active
        df = pd.read_excel(xlsx_file, index_col=0)
        # Step 1 调整列宽
        max_allow_width = 40
        encoding_type = 'gbk' # 中文在 utf-8 中是三个字节，在 gbk 中才是两个
        df_lens = pd.DataFrame()
        # 1.1 索引列
        indexs = df.index.copy()
        # 索引一般都是同类型可排序的数据集合，所以长度基本上可以用 max() 元素来代替
        indexs_max_length = len(str(indexs.max()).encode(encoding_type))
        index_width = max_allow_width if indexs_max_length > max_allow_width else indexs_max_length + 1
        df_lens = df_lens.append({'name':'索引列', '列号':get_column_letter(1),'列宽': int(index_width)}, ignore_index=True)
        # 1.2 数据列
        start_col = 2
        current_col = start_col
        for i in range(ws.max_column - 1):
            col_name = df.columns[i]
            date_max_length = df[col_name].apply(lambda x: len(str(x).encode(encoding_type))).max()
            data_width = max_allow_width if date_max_length > max_allow_width else date_max_length + 1
            # 如果值最大宽度不如标题，则应该至少把标题显示全
            data_width = data_width if len(col_name.encode(encoding_type)) + 1 < data_width else len(col_name.encode(encoding_type)) + 1
            df_lens = df_lens.append({'name': col_name, '列号':get_column_letter(current_col) ,'列宽': int(data_width)}, ignore_index=True)
            current_col += 1
        # 1.3 openpyxl 调整列宽
        for x in df_lens.itertuples():
            # print(x.列号, x.列宽)
            ws.column_dimensions[x.列号].width = x.列宽

        # Step 2 调整字体
        title_font = Font(name='Arial',
                        size=11,
                        bold=True,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='FF000000')
        value_font = Font(name='Arial',
                        size=11,
                        bold=False,
                        italic=False,
                        vertAlign=None,
                        underline='none',
                        strike=False,
                        color='FF000000')
        # 2.1 标题字体
        for row in ws.iter_rows(max_row=1):
            for cell in row:
                cell.font = title_font
        # 2.2 数据字体
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = value_font

        # Step 3 冻结首行
        ws.freeze_panes = 'A2'

        # Step 4 筛选
        ws.auto_filter.ref = "B:{0}".format(df_lens['列号'].values[-1])

        # Step 5 输出
        wb.save(xlsx_file)

if __name__ == "__main__":
    fmt = xlsxfmt()
    # fmt.format_xlsx_file('D:\github\deal-record-temp\output\康力泉_全部记录.xlsx')
    fmt.format_all_xlsx_files()